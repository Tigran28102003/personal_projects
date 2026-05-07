"""
build_dataset.py
Сборка панельного датасета для прогнозирования ВРП субъектов РФ.
Источник: Росстат «Социально-экономические показатели по субъектам РФ», 2025.
Выход: panel_dataset.csv, panel_dataset_no_lags.csv
"""

import re
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Агрегированные строки, которые нужно пропустить
# ---------------------------------------------------------------------------
SKIP_ROWS = {
    'центральный федеральный округ',
    'северо-западный федеральный округ',
    'южный федеральный округ',
    'северо-кавказский федеральный округ',
    'приволжский федеральный округ',
    'уральский федеральный округ',
    'сибирский федеральный округ',
    'дальневосточный федеральный округ',
    'крымский федеральный округ',
    'валовой региональный продукт по субъектам',
    'фактическое конечное потребление домашних хозяйств',
    'валовое накопление основного капитала',
    'среднегодовая численность занятых в экономике',
    'численность рабочей силы',
    'численность безработных',
    'уровень безработицы',
    'среднедушевые денежные доходы',
    'реальные доходы населения',
    'среднемесячная номинальная начисленная заработная плата',
    'основные фонды в экономике',
    'инвестиции в основной капитал',
    'объем отгруженных товаров',
    'оборот розничной торговли',
    'доходы консолидированных бюджетов',
    'расходы консолидированных бюджетов',
    'численность постоянного населения',
    'индекс потребительских цен',
    'перевозки грузов',
    'в том числе:',
}

AGGREGATE_REGION_REPLACEMENTS = {
    'архангельская область без автономного округа': 'архангельская область',
    'тюменская область (без ао)': 'тюменская область',
}

AGGREGATE_REGION_PAIRS = [
    ('архангельская область', 'архангельская область без автономного округа'),
    ('тюменская область', 'тюменская область (без ао)'),
]

CBR_KEY_RATE_URL = (
    'https://www.cbr.ru/eng/hd_base/KeyRate/'
    '?UniDbQuery.Posted=True&UniDbQuery.From=17.09.2013&UniDbQuery.To=06.05.2026'
)

CBR_REFINANCING_RATE_URL = 'https://www.cbr.ru/eng/statistics/idkp_br/refinancing_rates1/'
REGIONS_COLLECTION_DIR = DATA_DIR / 'data_regions_collection_102_v20260313_xlsx'
REGIONAL_CPI_GROWTH_PATH = DATA_DIR / 'reg_cpd.xlsx'
USD_RUB_PATH = DATA_DIR / 'RC_F01_01_2000_T07_05_2026.xlsx'

REFINANCING_RATE_FALLBACK = [
    ('1999-06-10', 55.00),
    ('2000-01-24', 45.00),
    ('2000-03-07', 38.00),
    ('2000-03-21', 33.00),
    ('2000-07-10', 28.00),
    ('2000-11-04', 25.00),
    ('2002-04-09', 23.00),
    ('2002-08-07', 21.00),
    ('2003-02-17', 18.00),
    ('2003-06-21', 16.00),
    ('2004-01-15', 14.00),
    ('2004-06-15', 13.00),
    ('2005-12-26', 12.00),
    ('2006-06-26', 11.50),
    ('2006-10-23', 11.00),
    ('2007-01-29', 10.50),
    ('2007-06-19', 10.00),
    ('2008-02-04', 10.25),
    ('2008-04-29', 10.50),
    ('2008-06-10', 10.75),
    ('2008-07-14', 11.00),
    ('2008-11-12', 12.00),
    ('2008-12-01', 13.00),
    ('2009-04-24', 12.50),
    ('2009-05-14', 12.00),
    ('2009-06-05', 11.50),
    ('2009-07-13', 11.00),
    ('2009-08-10', 10.75),
    ('2009-09-15', 10.50),
    ('2009-09-30', 10.00),
    ('2009-10-30', 9.50),
    ('2009-11-25', 9.00),
    ('2009-12-28', 8.75),
    ('2010-02-24', 8.50),
    ('2010-03-29', 8.25),
    ('2010-04-30', 8.00),
    ('2010-06-01', 7.75),
    ('2011-02-28', 8.00),
    ('2011-05-03', 8.25),
    ('2011-12-26', 8.00),
    ('2012-09-14', 8.25),
]

KEY_RATE_FALLBACK = [
    ('2013-09-17', 5.50),
]

MISSING_COLLECTION_VALUES = {-99999999, -77777777}

REGIONS_COLLECTION_FEATURES = [
    # Демография и структура населения: длинная история, низкая доля пропусков.
    {
        'filename': 'data_01_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110050',
        'feature': 'working_age_share',
    },
    {
        'filename': 'data_01_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110197',
        'feature': 'migration_growth_per_10k',
    },
    {
        'filename': 'data_01_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110256',
        'feature': 'life_expectancy',
    },
    {
        'filename': 'data_01_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110414',
        'feature': 'rural_population_share',
    },
    # Производственная структура.
    {
        'filename': 'data_12_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110116',
        'feature': 'mining_idx_collection',
    },
    {
        'filename': 'data_12_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110119',
        'feature': 'manufacturing_idx_collection',
    },
    {
        'filename': 'data_12_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110241',
        'feature': 'mining_output_collection',
    },
    {
        'filename': 'data_12_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110244',
        'feature': 'manufacturing_output_collection',
    },
    {
        'filename': 'data_13_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110342',
        'feature': 'agriculture_output_collection',
    },
    {
        'filename': 'data_14_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110017',
        'feature': 'housing_commissioning_per_1000',
    },
    # Услуги, торговля и транспорт.
    {
        'filename': 'data_15_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110218',
        'feature': 'public_catering',
    },
    {
        'filename': 'data_15_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110247',
        'feature': 'paid_services',
    },
    {
        'filename': 'data_15_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110248',
        'feature': 'paid_services_idx',
    },
    {
        'filename': 'data_16_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110061',
        'feature': 'road_freight_turnover',
    },
    {
        'filename': 'data_16_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110293',
        'feature': 'road_freight_shipments',
    },
    # Финансовая глубина региона и стоимость жизни.
    {
        'filename': 'data_19_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110097',
        'feature': 'household_ruble_credit_debt',
    },
    {
        'filename': 'data_19_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110099',
        'feature': 'mortgage_ruble_credit_debt',
    },
    {
        'filename': 'data_19_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110100',
        'feature': 'corporate_ruble_credit_debt',
    },
    {
        'filename': 'data_20_socio_economic_102_v20260313.xlsx',
        'indicator_code': 'Y477110393',
        'feature': 'min_food_basket_cost',
        'indicator_unit': 'Рублей',
    },
]

# ---------------------------------------------------------------------------
# Нормализация и алиасы названий регионов
# ---------------------------------------------------------------------------
REGION_ALIASES = {
    'г. москва': 'москва',
    'г.москва': 'москва',
    'город москва': 'москва',
    'г. санкт-петербург': 'санкт-петербург',
    'г.санкт-петербург': 'санкт-петербург',
    'город санкт-петербург': 'санкт-петербург',
    'г. севастополь': 'севастополь',
    'кемеровская область - кузбасс': 'кемеровская область',
    'ханты-мансийский автономный округ - югра': 'ханты-мансийский ао',
    'ханты-мансийский авт.округ - югра': 'ханты-мансийский ао',
    'ханты-мансийский автономный округ–югра': 'ханты-мансийский ао',
    'ханты-мансийский авт. округ - югра': 'ханты-мансийский ао',
    'ямало-ненецкий автономный округ': 'ямало-ненецкий ао',
    'ямало-ненецкий авт. округ': 'ямало-ненецкий ао',
    'чукотский автономный округ': 'чукотский ао',
    'чукотский авт. округ': 'чукотский ао',
    'ненецкий автономный округ': 'ненецкий ао',
    'ненецкий авт. округ': 'ненецкий ао',
    'еврейская автономная область': 'еврейская ао',
    'еврейская авт.область': 'еврейская ао',
    'еврейская авт. область': 'еврейская ао',
    'кабардино-балкарская республика': 'кабардино-балкария',
    'карачаево-черкесская республика': 'карачаево-черкесия',
    'республика северная осетия-алания': 'северная осетия',
    'республика северная осетия - алания': 'северная осетия',
    'чеченская республика': 'чечня',
    'республика адыгея': 'адыгея',
    'республика адыгея (адыгея)': 'адыгея',
    'республика алтай': 'республика алтай',
    'алтайский край': 'алтайский край',
    'республика башкортостан': 'башкортостан',
    'республика бурятия': 'бурятия',
    'республика дагестан': 'дагестан',
    'республика ингушетия': 'ингушетия',
    'республика калмыкия': 'калмыкия',
    'республика карелия': 'карелия',
    'республика коми': 'коми',
    'республика крым': 'крым',
    'республика марий эл': 'марий эл',
    'республика мордовия': 'мордовия',
    'республика саха (якутия)': 'якутия',
    'республика татарстан': 'татарстан',
    'республика тыва': 'тыва',
    'республика хакасия': 'хакасия',
    'удмуртская республика': 'удмуртия',
    'чувашская республика': 'чувашия',
    'тюменская область (без автономных округов)': 'тюменская область (без ао)',
    'тюменская область без автономных округов': 'тюменская область (без ао)',
    'тюменская область (кроме ханты-мансийского автономного округа-югры и ямало-ненецкого автономного округа)': 'тюменская область',
    'тюменская область (кроме ханты-мансийского автономного округа - югры и ямало-ненецкого автономного округа)': 'тюменская область',
    'архангельская область (кроме ненецкого автономного округа)': 'архангельская область',
    'пермский край': 'пермский край',
    'камчатский край': 'камчатский край',
    'забайкальский край': 'забайкальский край',
}


def normalize_region(name: str) -> str:
    if not isinstance(name, str):
        return ''
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    # заменить все виды тире/дефисов на обычный дефис для единообразия
    name = re.sub(r'[–—‒‐]', '-', name)
    name = name.lower()
    # убрать сноски вида "1)" "2)" в конце
    name = re.sub(r'\s*\d+\)\s*$', '', name).strip()
    return REGION_ALIASES.get(name, name)


def is_skip_row(name: str) -> bool:
    """Возвращает True, если строка — агрегат (федеральный округ или итоговая строка)."""
    n = name.lower().strip()
    for skip in SKIP_ROWS:
        if n.startswith(skip):
            return True
    if re.search(r'федеральный округ', n):
        return True
    if re.search(r'российская федерация', n):
        return True
    if re.search(r'по субъектам российской федерации', n):
        return True
    if n.startswith('в том числе'):
        return True
    if len(n) < 4:
        return True
    return False


def normalize_subject_panel(panel: pd.DataFrame) -> pd.DataFrame:
    out = panel.copy()
    out = out[~out['region'].str.startswith('в том числе', na=False)].copy()
    for aggregate, residual in AGGREGATE_REGION_PAIRS:
        has_aggregate = (out['region'] == aggregate).any()
        has_residual = (out['region'] == residual).any()
        if has_aggregate and has_residual:
            out = out[out['region'] != aggregate].copy()
    out['region'] = out['region'].replace(AGGREGATE_REGION_REPLACEMENTS)
    out = out.sort_values(['region', 'year'])
    out = out.drop_duplicates(subset=['region', 'year'], keep='first')
    return out.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Чтение листа: широкий → длинный формат
# ---------------------------------------------------------------------------
def read_rosstat_sheet(filepath: Path, sheetname: str, feature_name: str,
                       year_min: int = 2000, year_max: int = 2023) -> pd.DataFrame:
    """
    Читает лист Росстата (регионы × годы) и возвращает длинный DataFrame:
    columns: ['region', 'year', feature_name]
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [!] Не удалось открыть {filepath.name}: {e}")
        return pd.DataFrame(columns=['region', 'year', feature_name])

    if sheetname not in wb.sheetnames:
        print(f"  [!] Лист '{sheetname}' не найден в {filepath.name}")
        wb.close()
        return pd.DataFrame(columns=['region', 'year', feature_name])

    ws = wb[sheetname]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Найти строку-заголовок с годами
    header_idx = None
    year_cols = {}  # col_index → year
    for i, row in enumerate(rows):
        nums = [(j, v) for j, v in enumerate(row)
                if isinstance(v, (int, float)) and year_min <= int(v) <= year_max]
        if len(nums) >= 3:
            header_idx = i
            year_cols = {j: int(v) for j, v in nums}
            break

    if header_idx is None:
        print(f"  [!] Строка с годами не найдена: {filepath.name} [{sheetname}]")
        return pd.DataFrame(columns=['region', 'year', feature_name])

    records = []
    for row in rows[header_idx + 1:]:
        # Первый непустой текстовый элемент — название региона
        region_raw = None
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                region_raw = cell.strip()
                break
        if region_raw is None:
            continue

        region = normalize_region(region_raw)
        if not region or is_skip_row(region_raw):
            continue

        for col_idx, year in year_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, (int, float)):
                records.append({'region': region, 'year': year, feature_name: val})
            else:
                records.append({'region': region, 'year': year, feature_name: np.nan})

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Убрать дубликаты (иногда регион встречается дважды из-за сносок)
    df = df.drop_duplicates(subset=['region', 'year'], keep='first')
    return df


def load_regions_collection_features(
    collection_dir: Path = REGIONS_COLLECTION_DIR,
    year_min: int = 2000,
    year_max: int = 2023,
) -> pd.DataFrame:
    """
    Читает новую длинную коллекцию «Социально-экономические показатели
    регионов России» и возвращает wide-таблицу выбранных региональных
    показателей: region × year × feature.
    """
    if not collection_dir.exists():
        print(f"  [!] Новая коллекция региональных данных не найдена: {collection_dir}")
        return pd.DataFrame(columns=['region', 'year'])

    frames = []
    features_by_file: dict[str, list[dict]] = {}
    for spec in REGIONS_COLLECTION_FEATURES:
        features_by_file.setdefault(spec['filename'], []).append(spec)

    for filename, specs in features_by_file.items():
        path = collection_dir / filename
        if not path.exists():
            print(f"  [!] Файл новой коллекции не найден: {path.name}")
            continue
        codes = {spec['indicator_code'] for spec in specs}
        try:
            raw = pd.read_excel(
                path,
                sheet_name='Sheet1',
                usecols=[
                    'indicator_code',
                    'object_name',
                    'object_level',
                    'year',
                    'indicator_value',
                    'indicator_unit',
                    'version_date',
                ],
                dtype={'indicator_code': str},
            )
        except Exception as e:
            print(f"  [!] Не удалось прочитать новую коллекцию {path.name}: {e}")
            continue

        raw = raw[raw['indicator_code'].isin(codes)].copy()
        if raw.empty:
            continue
        level = raw['object_level'].astype(str).str.lower()
        raw = raw[level.str.contains('регион', na=False)].copy()
        raw['year'] = pd.to_numeric(raw['year'], errors='coerce')
        raw = raw[(raw['year'] >= year_min) & (raw['year'] <= year_max)].copy()
        raw['indicator_value'] = pd.to_numeric(raw['indicator_value'], errors='coerce')
        raw.loc[raw['indicator_value'].isin(MISSING_COLLECTION_VALUES), 'indicator_value'] = np.nan
        raw['region'] = raw['object_name'].map(normalize_region)
        raw = raw[(raw['region'] != '') & ~raw['object_name'].map(is_skip_row)].copy()

        for spec in specs:
            part = raw[raw['indicator_code'] == spec['indicator_code']].copy()
            unit = spec.get('indicator_unit')
            if unit is not None:
                part = part[part['indicator_unit'].astype(str) == unit].copy()
            if part.empty:
                continue
            part['version_date'] = pd.to_datetime(part['version_date'], errors='coerce')
            part = part.sort_values(['region', 'year', 'version_date'])
            part = part.drop_duplicates(subset=['region', 'year'], keep='last')
            part = part[['region', 'year', 'indicator_value']].rename(
                columns={'indicator_value': spec['feature']}
            )
            frames.append(part)

    if not frames:
        return pd.DataFrame(columns=['region', 'year'])

    out = frames[0]
    for frame in frames[1:]:
        out = out.merge(frame, on=['region', 'year'], how='outer')
    out = normalize_subject_panel(out)
    return out


def merge_collection_features(panel: pd.DataFrame, collection: pd.DataFrame) -> pd.DataFrame:
    if collection.empty:
        return panel
    out = panel.merge(collection, on=['region', 'year'], how='left')
    if 'agriculture' not in out.columns and 'agriculture_output_collection' in out.columns:
        out['agriculture'] = out['agriculture_output_collection']
    if 'mining_idx_collection' in out.columns:
        out['mining_index'] = out['mining_idx_collection']
    if 'manufacturing_idx_collection' in out.columns:
        out['manufacturing_index'] = out['manufacturing_idx_collection']
    if 'road_freight_shipments' in out.columns and 'freight_transport' in out.columns:
        out['freight_transport'] = out['freight_transport'].combine_first(out['road_freight_shipments'])
    return out


def load_regional_cpi_yoy(
    path: Path = REGIONAL_CPI_GROWTH_PATH,
    year_min: int = 2000,
    year_max: int = 2023,
) -> pd.DataFrame:
    """
    Читает первую страницу reg_cpd.xlsx: месячные темпы прироста
    потребительских цен год к году по регионам. Возвращает годовые
    средние и значения на конец года.
    """
    if not path.exists():
        print(f"  [!] Файл региональной инфляции не найден: {path}")
        return pd.DataFrame(columns=['region', 'year'])
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [!] Не удалось открыть файл региональной инфляции {path.name}: {e}")
        return pd.DataFrame(columns=['region', 'year'])

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame(columns=['region', 'year'])

    header = rows[0]
    date_cols = []
    for idx, value in enumerate(header[1:], start=1):
        date = pd.to_datetime(value, errors='coerce')
        if pd.notna(date):
            date_cols.append((idx, pd.Timestamp(date)))
    records = []
    for row in rows[1:]:
        region_raw = row[0] if row else None
        if not isinstance(region_raw, str):
            continue
        region = normalize_region(region_raw)
        if not region or region == 'россия' or is_skip_row(region_raw):
            continue
        for idx, date in date_cols:
            if date.year < year_min or date.year > year_max:
                continue
            value = row[idx] if idx < len(row) else None
            value = pd.to_numeric(value, errors='coerce')
            records.append(
                {
                    'region': region,
                    'year': int(date.year),
                    'date': date,
                    'regional_cpi_yoy_value': value,
                }
            )

    monthly = pd.DataFrame(records)
    if monthly.empty:
        return pd.DataFrame(columns=['region', 'year'])
    monthly = monthly.dropna(subset=['regional_cpi_yoy_value'])
    monthly = monthly.sort_values(['region', 'year', 'date'])
    annual = (
        monthly.groupby(['region', 'year'], as_index=False)
        .agg(
            regional_cpi_yoy=('regional_cpi_yoy_value', 'mean'),
            regional_cpi_yoy_eop=('regional_cpi_yoy_value', 'last'),
            regional_cpi_yoy_std=('regional_cpi_yoy_value', 'std'),
        )
    )
    annual['regional_cpi_yoy_change'] = annual.groupby('region')['regional_cpi_yoy_eop'].diff()
    annual = normalize_subject_panel(annual)
    return annual


def load_usd_rub_rates(
    path: Path = USD_RUB_PATH,
    year_min: int = 2000,
    year_max: int = 2023,
) -> pd.DataFrame:
    """
    Читает дневной курс доллар/рубль из файла ЦБ и агрегирует его до
    годовых макропоказателей.
    """
    if not path.exists():
        print(f"  [!] Файл курса USD/RUB не найден: {path}")
        return pd.DataFrame(columns=['year'])
    try:
        raw = pd.read_excel(path, sheet_name=0)
    except Exception as e:
        print(f"  [!] Не удалось прочитать курс USD/RUB {path.name}: {e}")
        return pd.DataFrame(columns=['year'])
    required = {'data', 'curs'}
    if not required.issubset(raw.columns):
        print(f"  [!] В файле курса USD/RUB нет колонок {required}: {path.name}")
        return pd.DataFrame(columns=['year'])
    raw = raw.copy()
    raw['date'] = pd.to_datetime(raw['data'], errors='coerce')
    raw['usd_rub_value'] = pd.to_numeric(raw['curs'], errors='coerce')
    raw = raw.dropna(subset=['date', 'usd_rub_value'])
    raw = raw[(raw['date'].dt.year >= year_min) & (raw['date'].dt.year <= year_max)]
    if raw.empty:
        return pd.DataFrame(columns=['year'])
    raw = raw.sort_values('date')
    annual = (
        raw.groupby(raw['date'].dt.year)
        .agg(
            usd_rub=('usd_rub_value', 'mean'),
            usd_rub_eop=('usd_rub_value', 'last'),
            usd_rub_min=('usd_rub_value', 'min'),
            usd_rub_max=('usd_rub_value', 'max'),
            usd_rub_volatility=('usd_rub_value', 'std'),
        )
        .reset_index()
        .rename(columns={'date': 'year'})
    )
    annual['year'] = annual['year'].astype(int)
    annual['usd_rub_change'] = annual['usd_rub_eop'].diff()
    annual['usd_rub_pct_change'] = annual['usd_rub_eop'].pct_change(fill_method=None) * 100
    annual['usd_rub_range'] = annual['usd_rub_max'] - annual['usd_rub_min']
    return annual


# ---------------------------------------------------------------------------
# Классификация регионов
# ---------------------------------------------------------------------------
REGION_TYPE_MAP = {
    # resource: сырьевые/добывающие
    'ханты-мансийский ао': 'resource',
    'ямало-ненецкий ао': 'resource',
    'тюменская область (без ао)': 'resource',
    'тюменская область': 'resource',
    'сахалинская область': 'resource',
    'кемеровская область': 'resource',
    'коми': 'resource',
    'ненецкий ао': 'resource',
    'якутия': 'resource',
    'иркутская область': 'resource',
    'оренбургская область': 'resource',
    'астраханская область': 'resource',
    'томская область': 'resource',
    'магаданская область': 'resource',
    'чукотский ао': 'resource',
    'мурманская область': 'resource',
    # service: финансово-сервисные центры
    'москва': 'service',
    'санкт-петербург': 'service',
    'московская область': 'service',
    # industrial: обрабатывающая промышленность
    'свердловская область': 'industrial',
    'челябинская область': 'industrial',
    'самарская область': 'industrial',
    'нижегородская область': 'industrial',
    'тульская область': 'industrial',
    'липецкая область': 'industrial',
    'вологодская область': 'industrial',
    'ярославская область': 'industrial',
    'владимирская область': 'industrial',
    'тверская область': 'industrial',
    'калужская область': 'industrial',
    'пермский край': 'industrial',
    'удмуртия': 'industrial',
    'татарстан': 'industrial',
    'башкортостан': 'industrial',
    'омская область': 'industrial',
    'новосибирская область': 'industrial',
    'красноярский край': 'industrial',
    'хабаровский край': 'industrial',
    'приморский край': 'industrial',
    # agro: аграрные
    'краснодарский край': 'agro',
    'ставропольский край': 'agro',
    'ростовская область': 'agro',
    'белгородская область': 'agro',
    'воронежская область': 'agro',
    'курская область': 'agro',
    'тамбовская область': 'agro',
    'пензенская область': 'agro',
    'саратовская область': 'agro',
    'волгоградская область': 'agro',
    'алтайский край': 'agro',
    'адыгея': 'agro',
    'калмыкия': 'agro',
}


def get_region_type(region: str) -> str:
    return REGION_TYPE_MAP.get(region, 'mixed')


def load_cbr_policy_rates(year_min: int | None = None, year_max: int | None = None) -> pd.DataFrame:
    year_min = YEAR_MIN if year_min is None else year_min
    year_max = YEAR_MAX if year_max is None else year_max
    key_changes = load_key_rate_changes()
    refinancing_changes = load_refinancing_rate_changes()
    return annualize_policy_rates(key_changes, refinancing_changes, year_min, year_max)


def load_key_rate_changes() -> pd.DataFrame:
    web = read_key_rate_from_web()
    local = read_key_rate_from_excel(DATA_DIR / 'Процентные ставки Банк России.xlsx')
    fallback = pd.DataFrame(KEY_RATE_FALLBACK, columns=['date', 'key_rate'])
    frames = [df for df in [fallback, web, local] if df is not None and not df.empty]
    if not frames:
        return pd.DataFrame(columns=['date', 'key_rate'])
    out = pd.concat(frames, ignore_index=True)
    out['date'] = pd.to_datetime(out['date'], errors='coerce')
    out['key_rate'] = out['key_rate'].map(coerce_rate)
    out = out.dropna(subset=['date', 'key_rate'])
    out = out.sort_values('date').drop_duplicates(subset=['date'], keep='last')
    return out.reset_index(drop=True)


def read_key_rate_from_excel(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=['date', 'key_rate'])
    try:
        raw = pd.read_excel(path, sheet_name='таб_дни изменений', header=None, skiprows=6)
    except Exception as e:
        print(f"  [!] Не удалось прочитать локальный файл ставок ЦБ: {e}")
        return pd.DataFrame(columns=['date', 'key_rate'])
    if raw.shape[1] < 2:
        return pd.DataFrame(columns=['date', 'key_rate'])
    out = raw[[0, 1]].copy()
    out.columns = ['date', 'key_rate']
    out['date'] = pd.to_datetime(out['date'], errors='coerce')
    out['key_rate'] = out['key_rate'].map(coerce_rate)
    return out.dropna(subset=['date', 'key_rate'])


def read_key_rate_from_web() -> pd.DataFrame:
    try:
        tables = pd.read_html(CBR_KEY_RATE_URL)
    except Exception as e:
        print(f"  [!] Не удалось загрузить ключевую ставку ЦБ с сайта: {e}")
        return pd.DataFrame(columns=['date', 'key_rate'])
    for table in tables:
        if table.shape[1] < 2:
            continue
        table = table.copy()
        cols = [str(c).lower() for c in table.columns]
        date_col = next((table.columns[i] for i, c in enumerate(cols) if 'date' in c or 'дата' in c), table.columns[0])
        rate_col = next((table.columns[i] for i, c in enumerate(cols) if 'rate' in c or 'став' in c), table.columns[1])
        out = table[[date_col, rate_col]].copy()
        out.columns = ['date', 'key_rate']
        out['date'] = pd.to_datetime(out['date'], errors='coerce', dayfirst=True)
        out['key_rate'] = out['key_rate'].map(coerce_rate)
        out = out.dropna(subset=['date', 'key_rate'])
        if not out.empty:
            return out
    return pd.DataFrame(columns=['date', 'key_rate'])


def load_refinancing_rate_changes() -> pd.DataFrame:
    web = read_refinancing_rate_from_web()
    fallback = pd.DataFrame(REFINANCING_RATE_FALLBACK, columns=['date', 'refinancing_rate'])
    frames = [df for df in [fallback, web] if df is not None and not df.empty]
    if not frames:
        return pd.DataFrame(columns=['date', 'refinancing_rate'])
    out = pd.concat(frames, ignore_index=True)
    out['date'] = pd.to_datetime(out['date'], errors='coerce')
    out['refinancing_rate'] = out['refinancing_rate'].map(coerce_rate)
    out = out.dropna(subset=['date', 'refinancing_rate'])
    out = out.sort_values('date').drop_duplicates(subset=['date'], keep='last')
    return out.reset_index(drop=True)


def read_refinancing_rate_from_web() -> pd.DataFrame:
    try:
        tables = pd.read_html(CBR_REFINANCING_RATE_URL)
    except Exception as e:
        print(f"  [!] Не удалось загрузить ставку рефинансирования ЦБ с сайта: {e}")
        return pd.DataFrame(columns=['date', 'refinancing_rate'])
    records = []
    for table in tables:
        if table.empty:
            continue
        for _, row in table.iterrows():
            text = ' '.join(str(v) for v in row.to_list() if pd.notna(v))
            start = parse_refinancing_start_date(text)
            rate = coerce_rate(row.iloc[-1])
            if start is not None and rate is not None:
                records.append((start, rate))
    return pd.DataFrame(records, columns=['date', 'refinancing_rate'])


def parse_refinancing_start_date(text: str):
    match = re.search(
        r'([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})|'
        r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})',
        text,
    )
    if not match:
        return None
    if match.group(1):
        date_text = f"{match.group(1)} {match.group(2)}, {match.group(3)}"
    else:
        date_text = f"{match.group(5)} {match.group(4)}, {match.group(6)}"
    return pd.to_datetime(date_text, errors='coerce')


def annualize_policy_rates(
    key_changes: pd.DataFrame,
    refinancing_changes: pd.DataFrame,
    year_min: int,
    year_max: int,
) -> pd.DataFrame:
    dates = pd.date_range(f'{year_min}-01-01', f'{year_max}-12-31', freq='D')
    daily = pd.DataFrame(index=dates)
    daily['key_rate'] = build_daily_rate_series(key_changes, 'key_rate', dates)
    daily['refinancing_rate'] = build_daily_rate_series(
        refinancing_changes, 'refinancing_rate', dates
    )
    daily.loc[daily.index >= pd.Timestamp('2016-01-01'), 'refinancing_rate'] = np.nan
    daily['policy_rate'] = daily['key_rate'].combine_first(daily['refinancing_rate'])

    avg = daily.groupby(daily.index.year)[['key_rate', 'refinancing_rate', 'policy_rate']].mean()
    eop = daily.groupby(daily.index.year)[['policy_rate']].last().rename(
        columns={'policy_rate': 'policy_rate_eop'}
    )
    annual = avg.join(eop)
    annual['policy_rate_change'] = annual['policy_rate_eop'].diff()
    annual = annual.reset_index().rename(columns={'index': 'year'})
    annual['year'] = annual['year'].astype(int)
    return annual


def build_daily_rate_series(changes: pd.DataFrame, rate_col: str, dates: pd.DatetimeIndex) -> pd.Series:
    series = pd.Series(index=dates, dtype=float)
    if changes.empty:
        return series
    work = changes.copy()
    work['date'] = pd.to_datetime(work['date'], errors='coerce')
    work[rate_col] = work[rate_col].map(coerce_rate)
    work = work.dropna(subset=['date', rate_col]).sort_values('date')
    work = work[work['date'] <= dates.max()]
    if work.empty:
        return series
    all_dates = pd.date_range(min(work['date'].min(), dates.min()), dates.max(), freq='D')
    full = pd.Series(index=all_dates, dtype=float)
    full.loc[work['date']] = work[rate_col].to_numpy()
    full = full.sort_index().ffill()
    return full.reindex(dates)


def coerce_rate(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.replace(',', '.')
        match = re.search(r'-?\d+(?:\.\d+)?', value)
        if not match:
            return None
        value = match.group(0)
    try:
        return float(value)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Источники данных
# ---------------------------------------------------------------------------
SOURCES = [
    # (filename, sheetname, feature_name)
    ('Раздел 1 - Население.xlsx',                        '1.1.',     'population'),
    ('Раздел 2 - Труд.xlsx',                             '2.3.',     'employed'),
    ('Раздел 2 - Труд.xlsx',                             '2.10.1.',  'unemployment_rate'),
    ('Раздел 3 - Уровень жизни населения.xlsx',          '3.1.1.',   'real_income_idx'),
    ('Раздел 3 - Уровень жизни населения.xlsx',          '3.2.',     'avg_income_percapita'),
    ('Раздел 3 - Уровень жизни населения.xlsx',          '3.4.',     'avg_wage'),
    ('Раздел 9 - Основные фонды.xlsx',                   '9.1.',     'fixed_assets'),
    ('Раздел 10 - Инвестиции.xlsx',                      '10.1.',    'investment'),
    ('Раздел 10 - Инвестиции.xlsx',                      '10.3.',    'investment_idx'),
    ('Раздел 12 - Промышленное производство.xlsx',       '12.2.2.',  'mining_output'),
    ('Раздел 12 - Промышленное производство.xlsx',       '12.2.4.',  'manufacturing_output'),
    ('Раздел 14 - Строительство.xlsx',                   '14.1.',    'construction'),
    ('Раздел 15 - Торговля и услуги населению.xlsx',     '15.1.',    'retail_trade'),
    ('Раздел 16 - Транспорт.xlsx',                       '16.3.',    'freight_transport'),
    ('Раздел 18 - Наука и инновации.xlsx',               '18.1.',    'rd_expenditure'),
    ('Раздел 19 - Финансы.xlsx',                         '19.1.',    'budget_revenue'),
    ('Раздел 19 - Финансы.xlsx',                         '19.2.',    'budget_expenditure'),
    ('Раздел 20 - Цены и тарифы.xlsx',                   '20.1.',    'cpi'),
    # Целевые переменные
    ('Раздел 8 - Валовой региональный продукт.xlsx',     '8.1.',     'vrp'),
    ('Раздел 8 - Валовой региональный продукт.xlsx',     '8.2.',     'vrp_per_capita'),
    ('Раздел 8 - Валовой региональный продукт.xlsx',     '8.3.',     'vrp_idx'),
]

YEAR_MIN, YEAR_MAX = 2000, 2023


# ---------------------------------------------------------------------------
# Основная функция
# ---------------------------------------------------------------------------
def build_panel() -> pd.DataFrame:
    # 1. Загрузить все источники
    all_dfs = {}
    for filename, sheetname, feature in SOURCES:
        print(f"Читаю {filename} [{sheetname}] → {feature}")
        df = read_rosstat_sheet(DATA_DIR / filename, sheetname, feature,
                                YEAR_MIN, YEAR_MAX)
        all_dfs[feature] = df

    # 2. Построить базу (регионы × годы) из целевой переменной vrp
    vrp_df = all_dfs['vrp']
    base_regions = sorted(vrp_df['region'].unique())
    base_years = list(range(YEAR_MIN, YEAR_MAX + 1))
    base = pd.DataFrame(
        [(r, y) for r in base_regions for y in base_years],
        columns=['region', 'year']
    )
    print(f"\nБазовые регионы: {len(base_regions)}, годы: {YEAR_MIN}–{YEAR_MAX}")

    # 3. Последовательный merge всех признаков
    panel = base.copy()
    for feature, df in all_dfs.items():
        if df.empty:
            panel[feature] = np.nan
            continue
        panel = panel.merge(df[['region', 'year', feature]], on=['region', 'year'], how='left')

    # 4. Убрать агрегаты и привести остаточные области без АО к субъектам РФ
    panel = normalize_subject_panel(panel)

    # 5. Новая длинная коллекция региональных показателей
    print("\nЧитаю новую коллекцию региональных социально-экономических показателей")
    collection = load_regions_collection_features(REGIONS_COLLECTION_DIR, YEAR_MIN, YEAR_MAX)
    panel = merge_collection_features(panel, collection)

    # 6. Региональная инфляция и курс USD/RUB
    regional_cpi = load_regional_cpi_yoy(REGIONAL_CPI_GROWTH_PATH, YEAR_MIN, YEAR_MAX)
    if not regional_cpi.empty:
        panel = panel.merge(regional_cpi, on=['region', 'year'], how='left')
        if 'cpi' in panel.columns:
            panel['cpi_from_regional_yoy'] = panel['regional_cpi_yoy'] + 100
            panel['cpi'] = panel['cpi'].combine_first(panel['cpi_from_regional_yoy'])

    usd_rub = load_usd_rub_rates(USD_RUB_PATH, YEAR_MIN, YEAR_MAX)
    if not usd_rub.empty:
        panel = panel.merge(usd_rub, on='year', how='left')

    # 7. Макропроцентные ставки ЦБ РФ
    rates = load_cbr_policy_rates(YEAR_MIN, YEAR_MAX)
    panel = panel.merge(rates, on='year', how='left')
    if 'cpi' in panel.columns:
        inflation = panel['cpi'] - 100
        if 'regional_cpi_yoy' in panel.columns:
            inflation = panel['regional_cpi_yoy'].combine_first(inflation)
        panel['policy_rate_real'] = panel['policy_rate'] - inflation

    # 8. Тип региона
    panel['region_type'] = panel['region'].map(get_region_type)

    # 9. Кризисные переменные
    panel['crisis_2008'] = panel['year'].isin([2008, 2009]).astype(int)
    panel['crisis_2014'] = panel['year'].isin([2014, 2015]).astype(int)
    panel['crisis_2020'] = panel['year'].isin([2020]).astype(int)
    panel['crisis_2022'] = panel['year'].isin([2022, 2023]).astype(int)
    panel['crisis_any'] = panel[['crisis_2008', 'crisis_2014', 'crisis_2020', 'crisis_2022']].max(axis=1)

    panel = panel.sort_values(['region', 'year']).reset_index(drop=True)
    panel['cumulative_crisis_years'] = panel.groupby('region')['crisis_any'].cumsum()

    # Взаимодействия тип_региона × кризис
    for crisis_col in ['crisis_2008', 'crisis_2014', 'crisis_2020', 'crisis_2022']:
        for rtype in ['resource', 'industrial', 'agro', 'service']:
            col = f'{rtype}_{crisis_col}'
            panel[col] = ((panel['region_type'] == rtype) & (panel[crisis_col] == 1)).astype(int)

    # 10. Лаговые признаки
    for lag in [1, 2, 3]:
        panel[f'vrp_lag{lag}'] = panel.groupby('region')['vrp'].shift(lag)
        panel[f'vrp_per_capita_lag{lag}'] = panel.groupby('region')['vrp_per_capita'].shift(lag)

    lag1_features = ['investment', 'fixed_assets', 'employed', 'budget_revenue',
                     'construction', 'retail_trade', 'rd_expenditure', 'avg_wage',
                     'unemployment_rate', 'key_rate', 'refinancing_rate',
                     'policy_rate', 'policy_rate_eop', 'policy_rate_change',
                     'policy_rate_real', 'working_age_share',
                     'migration_growth_per_10k', 'life_expectancy',
                     'rural_population_share', 'mining_idx_collection',
                     'manufacturing_idx_collection', 'agriculture_output_collection',
                     'housing_commissioning_per_1000', 'public_catering',
                     'paid_services', 'paid_services_idx', 'road_freight_turnover',
                     'road_freight_shipments', 'household_ruble_credit_debt',
                     'mortgage_ruble_credit_debt', 'corporate_ruble_credit_debt',
                     'min_food_basket_cost', 'regional_cpi_yoy',
                     'regional_cpi_yoy_eop', 'regional_cpi_yoy_std',
                     'regional_cpi_yoy_change', 'usd_rub', 'usd_rub_eop',
                     'usd_rub_change', 'usd_rub_pct_change',
                     'usd_rub_volatility', 'usd_rub_range']
    for feat in lag1_features:
        if feat in panel.columns:
            panel[f'{feat}_lag1'] = panel.groupby('region')[feat].shift(1)

    # 11. Темпы прироста
    for feat in ['vrp', 'investment', 'population', 'fixed_assets', 'avg_wage']:
        if feat in panel.columns:
            panel[f'{feat}_growth'] = panel.groupby('region')[feat].pct_change(fill_method=None) * 100

    # 12. Производные признаки
    if 'investment' in panel.columns and 'vrp' in panel.columns:
        panel['investment_to_vrp'] = panel['investment'] / panel['vrp']
    if 'employed' in panel.columns and 'vrp' in panel.columns:
        panel['labor_productivity'] = panel['vrp'] / panel['employed']
    if 'budget_revenue' in panel.columns and 'vrp' in panel.columns:
        panel['budget_to_vrp'] = panel['budget_revenue'] / panel['vrp']

    # 13. Логарифмы уровневых переменных
    log_cols = ['vrp', 'vrp_per_capita', 'investment', 'fixed_assets', 'population',
                'budget_revenue', 'retail_trade', 'avg_wage', 'avg_income_percapita',
                'agriculture_output_collection', 'housing_commissioning_per_1000',
                'public_catering', 'paid_services', 'road_freight_turnover',
                'road_freight_shipments', 'household_ruble_credit_debt',
                'mortgage_ruble_credit_debt', 'corporate_ruble_credit_debt',
                'min_food_basket_cost']
    for col in log_cols:
        if col in panel.columns:
            panel[f'log_{col}'] = np.log1p(panel[col].clip(lower=0))

    # 14. Порядковый ID региона (для tree-моделей)
    panel['region_id'] = pd.Categorical(panel['region']).codes

    # Упорядочить столбцы
    id_cols = ['region', 'region_id', 'region_type', 'year']
    target_cols = ['vrp', 'vrp_per_capita', 'vrp_idx']
    other_cols = [c for c in panel.columns if c not in id_cols + target_cols]
    panel = panel[id_cols + target_cols + other_cols]

    return panel


def print_summary(panel: pd.DataFrame):
    print("\n" + "=" * 60)
    print(f"Итоговый датасет: {panel.shape[0]} строк × {panel.shape[1]} колонок")
    print(f"Регионов: {panel['region'].nunique()}")
    print(f"Лет: {panel['year'].min()}–{panel['year'].max()}")
    print(f"Тип региона: {panel['region_type'].value_counts().to_dict()}")
    print("\nПропуски по ключевым колонкам (%):")
    key_cols = ['vrp', 'vrp_per_capita', 'vrp_idx', 'population', 'employed',
                'investment', 'fixed_assets', 'avg_wage', 'budget_revenue',
                'retail_trade', 'cpi', 'mining_output', 'manufacturing_output',
                'policy_rate', 'key_rate', 'refinancing_rate',
                'regional_cpi_yoy', 'regional_cpi_yoy_eop', 'usd_rub',
                'usd_rub_eop', 'usd_rub_volatility',
                'working_age_share', 'life_expectancy', 'agriculture_output_collection',
                'paid_services', 'road_freight_turnover',
                'household_ruble_credit_debt', 'min_food_basket_cost']
    for col in key_cols:
        if col in panel.columns:
            pct = panel[col].isnull().mean() * 100
            print(f"  {col:<30} {pct:.1f}%")
    print("=" * 60)


if __name__ == '__main__':
    panel = build_panel()
    print_summary(panel)

    # Сохранить основной датасет
    out_main = DATA_DIR / 'panel_dataset.csv'
    panel.to_csv(out_main, index=False, encoding='utf-8-sig')
    print(f"\nСохранено: {out_main}")

    # Сохранить версию без лагов и производных (чистые уровни для EDA)
    no_lag_cols = [c for c in panel.columns
                   if not any(c.endswith(s) for s in ['_lag1', '_lag2', '_lag3',
                                                       '_growth', 'log_',
                                                       'labor_productivity',
                                                       'investment_to_vrp',
                                                       'budget_to_vrp'])
                   and not c.startswith('log_')]
    panel_eda = panel[no_lag_cols]
    out_eda = DATA_DIR / 'panel_dataset_no_lags.csv'
    panel_eda.to_csv(out_eda, index=False, encoding='utf-8-sig')
    print(f"Сохранено: {out_eda}")
